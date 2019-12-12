# -*- coding: utf-8 -*-

"""
Module implementing Form.
"""
from PyQt5.QtGui import QIcon, QPixmap, QPainter
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QWidget, QApplication, QMessageBox
from openpyxl import load_workbook
from Ui_FenbuhanshuNihe import Ui_FenbuhanshuNihe_Form
from scipy import stats
from scipy.stats import chi2
import math

class FenbuhanshuNihe_Form(QWidget, Ui_FenbuhanshuNihe_Form):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(FenbuhanshuNihe_Form, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('./image/icon.png')) 
    
    def paintEvent(self,event):
        painter = QPainter(self)
        pixmap = QPixmap("./image/background.png")
        #绘制窗口背景，平铺到整个窗口，随着窗口改变而改变
        painter.drawPixmap(self.rect(),pixmap) 
    
    @pyqtSlot()
    def on_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        try:
            a=self.lineEdit.text()
            book = load_workbook( filename=a )
            sheet = book.get_sheet_by_name("Sheet1")
            data=[]
            row_num=1
            while True:
                if sheet.cell(row=row_num, column=1).value == None:
                    break
                data.append(sheet.cell(row=row_num, column=1).value)
                row_num = row_num + 1
            
            parameter=stats.shapiro(data) 
            self.lineEdit_2.setText(str("{:.4f}".format(parameter[1])))
        
            if parameter[1]>0.05:  #检验水平规定为0.05
                self.lineEdit_4.setText(str("服从"))
            else:
                self.lineEdit_4.setText(str("不服从"))
            print(data)
            
        except :
            QMessageBox.information(self, "标题", "原文件路径输入错误", QMessageBox.Cancel)
            
    
    
    @pyqtSlot()
    def on_pushButton_2_clicked(self):
        """
        Slot documentation goes here.
        """
        try:
            a=self.lineEdit.text()
            book = load_workbook( filename=a )
            sheet = book.get_sheet_by_name("Sheet1")
            data_i=[]
            data_vi=[]
            row_num=1
            while True:
                if sheet.cell(row=row_num, column=1).value == None or sheet.cell(row=row_num, column=2).value == None:
                    break
                data_i.append(sheet.cell(row=row_num, column=1).value)
                data_vi.append(sheet.cell(row=row_num, column=2).value)
                row_num = row_num + 1
            
            print(data_i)
            print(data_vi)
        
            m=len(data_i)
            n=sum(data_vi)
        
            print(m)
            print(n)
        
            vi_2=[]
            H=0
            for i in range(0, m):
                H+=data_i[i]*data_vi[i]
                vi_2.append(data_vi[i]**2)
            lamda=H/n
        
            print(lamda)
            print(vi_2)
            vi_2_np=[]     #vi的平方除以n*pi
            for i in range(0, m):
                vi_2_np.append(vi_2[i]/(n*math.pow(lamda, data_i[i])*math.exp(-lamda)/math.factorial(data_i[i])))
        
            x=sum(vi_2_np)-n   #统计量
            print (x)
            p=chi2.sf(x, m-2)
            self.lineEdit_3.setText(str("{:.4f}".format(p)))
            if p>0.05:  #检验水平规定为0.05
                self.lineEdit_5.setText(str("服从"))
            else:
                self.lineEdit_5.setText(str("不服从"))
           
        except:
            QMessageBox.information(self, "标题", "原文件路径输入错误", QMessageBox.Cancel)
        
        
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = FenbuhanshuNihe_Form()
    ui.show()
    sys.exit(app.exec_())
