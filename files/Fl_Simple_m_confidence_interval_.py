# -*- coding: utf-8 -*-

"""
Module implementing Simple_m_confidence_interval_Form.
"""
from openpyxl import load_workbook
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import *
from Ui_Simple_m_confidence_interval_ import Ui_Simple_m_confidence_interval_Form
from scipy.stats import chi2, t, norm
from PyQt5.QtGui import QIcon, QPixmap, QPainter 
import numpy as np


class Simple_m_confidence_interval_Form(QWidget, Ui_Simple_m_confidence_interval_Form):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(Simple_m_confidence_interval_Form, self).__init__(parent)
        self.setupUi(self)
        self.widget.setVisible(False)
        self.setWindowIcon(QIcon('./image/icon.png')) 
        self.data = None
        
    def paintEvent(self,event):
        painter = QPainter(self)
        pixmap = QPixmap("./image/background.png")
        #绘制窗口背景，平铺到整个窗口，随着窗口改变而改变
        painter.drawPixmap(self.rect(),pixmap)
    
    @pyqtSlot()
    def on_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        m=float(self.doubleSpinBox_2.text())
        n=float(self.doubleSpinBox_3.text())
        if self.data == None:
            N=int(self.spinBox.text())
            x = np.random.normal(m, n, N)
        else:
            N = len(self.data)
            x = np.array(self.data)
        a=float(self.doubleSpinBox.text())
        mean, std = x.mean(), x.std(ddof=1)
        self.lineEdit_22.setText(str(mean))
        self.lineEdit_24.setText(str(std))
        y=norm.ppf(1-a/2,0,1)
        self.lineEdit_3.setText(str(round(y, 3)))
        w=mean-(y*n/(N**0.5))
        v=mean+(y*n/(N**0.5))
        conf_intveral1=(round(w, 3), round(v, 3))
        self.lineEdit_21.setText(str(conf_intveral1))
        self.lineEdit_4.setText('')
        self.lineEdit_5.setText('')
        self.lineEdit_25.setText('')
        if self.radioButton.isChecked():
            self.widget.setVisible(True)
            self.widget.mpl.draw_hist2(N, x, mean, std, m, n)
        elif self.radioButton_2.isChecked():
            self.widget.setVisible(True)
            self.widget.mpl.draw_hist(N, x)
        # TODO: not implemented yet
    
    @pyqtSlot()
    def on_pushButton_2_clicked(self):
        """
        Slot documentation goes here.
        """
        m=float(self.doubleSpinBox_2.text())
        n=float(self.doubleSpinBox_3.text())
        if self.data == None:
            N=int(self.spinBox.text())
            x = np.random.normal(m, n, N)
        else:
            N = len(self.data)
            x = np.array(self.data)
        a=float(self.doubleSpinBox.text())
        mean, std = x.mean(), x.std(ddof=1)
        self.lineEdit_22.setText(str(mean))
        self.lineEdit_24.setText(str(std))
        y=t.ppf(1-a/2, N-1)
        self.lineEdit_3.setText(str(round(y, 3)))
        e=mean-(y*std/((N-1)**0.5))
        f=mean+(y*std/((N-1)**0.5))
        conf_intveral1=(round(e, 3),  round(f, 3))
        self.lineEdit_21.setText(str(conf_intveral1))
        self.lineEdit_4.setText('')
        self.lineEdit_5.setText('')
        self.lineEdit_25.setText('')
        if self.radioButton.isChecked():
            self.widget.setVisible(True)
            self.widget.mpl.draw_hist2(N, x, mean, std, m, n)
        elif self.radioButton_2.isChecked():
            self.widget.setVisible(True)
            self.widget.mpl.draw_hist(N, x)
        # TODO: not implemented yet
    
    @pyqtSlot()
    def on_pushButton_3_clicked(self):
        """
        Slot documentation goes here.
        """
        m=float(self.doubleSpinBox_2.text())
        n=float(self.doubleSpinBox_3.text())
        if self.data == None:
            N=int(self.spinBox.text())
            x = np.random.normal(m, n, N)
        else:
            N = len(self.data)
            x = np.array(self.data)
        a=float(self.doubleSpinBox.text())
        mean, std = x.mean(), x.std(ddof=1)
        self.lineEdit_22.setText(str(mean))
        self.lineEdit_24.setText(str(std))
        self.lineEdit_3.setText('')
        self.lineEdit_21.setText('')
        p=chi2.ppf(a/2, N-1)
        q=chi2.ppf(1-a/2, N-1)
        self.lineEdit_4.setText(str(round(p, 3)))
        self.lineEdit_5.setText(str(round(q, 3)))
        conf_intveral2=(round((N*std)/q, 3), round((N*std)/p, 3))
        self.lineEdit_25.setText(str(conf_intveral2))
        if self.radioButton.isChecked():
            self.widget.setVisible(True)
            self.widget.mpl.draw_hist2(N, x, mean, std, m, n)
        elif self.radioButton_2.isChecked():
            self.widget.setVisible(True)
            self.widget.mpl.draw_hist(N, x)
        # TODO: not implemented yet
    
    @pyqtSlot()
    def on_pushButton_4_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        try:
            a=self.lineEdit.text()
            book = load_workbook(filename=a + ".xlsx")
            sheet = book.get_sheet_by_name("Sheet1")
            data=[]
            row_num=1
            while True:
                if sheet.cell(row=row_num, column=1).value == None:
                    break
                data.append(sheet.cell(row=row_num, column=1).value)
                row_num = row_num + 1
        except:
            QMessageBox.information(self, "标题", "请输入正确路径！", QMessageBox.Cancel)
        else:
            QMessageBox.information(self, "标题", "导入成功！", QMessageBox.Cancel)
        self.data = data
    

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = Simple_m_confidence_interval_Form()
    ui.show()
    sys.exit(app.exec_())  
