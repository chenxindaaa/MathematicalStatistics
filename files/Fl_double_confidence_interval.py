# -*- coding: utf-8 -*-

"""
Module implementing double_confidence_interval_Form.
"""
from openpyxl import load_workbook
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import *
from PyQt5.QtGui import QIcon, QPixmap, QPainter 
from Ui_double_confidence_interval import Ui_double_confidence_interval_Form
from scipy.stats import t, norm
import numpy as np


class double_confidence_interval_Form(QWidget, Ui_double_confidence_interval_Form):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(double_confidence_interval_Form, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('./image/icon.png')) 
        self.widget.setVisible(False)
        self.data = None
    
    def paintEvent(self,event):
        painter = QPainter(self)
        pixmap = QPixmap("./image/background.png")
        #绘制窗口背景，平铺到整个窗口，随着窗口改变而改变
        painter.drawPixmap(self.rect(),pixmap)
    
    @pyqtSlot()
    def on_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        if self.comboBox.currentText() == "均值差（方差已知）":
            m1=float(self.doubleSpinBox_2.text())
            n1=float(self.doubleSpinBox_3.text())
            m2=float(self.doubleSpinBox_4.text())
            n2=float(self.doubleSpinBox_5.text())
            if self.data == None:
                N1=int(self.spinBox.text())
                x = np.random.normal(m1, n1, N1)
                N2=int(self.spinBox_2.text())
                y = np.random.normal(m2, n2, N2)
            else:
                x = np.array(self.data[0])
                y = np.array(self.data[1])
                N1 = len(x)
                N2 = len(y)
            mean1, std1 = x.mean(), x.std(ddof=1)
            mean2, std2 = y.mean(), y.std(ddof=1)
            self.lineEdit_3.setText(str(round(mean1, 3)))
            self.lineEdit_4.setText(str(round(std1, 3)))
            self.lineEdit_5.setText(str(round(mean2, 3)))
            self.lineEdit_6.setText(str(round(std2, 3)))
            a=float(self.doubleSpinBox.text())
            z=norm.ppf(1-a/2,0,1)
            e=mean1-mean2-(z*((std1/N1+std2/N2)**0.5))
            f=mean1-mean2+(z*((std1/N1+std2/N2)**0.5))
            conf_intveral1=(round(e, 3), round(f, 3))
            self.lineEdit_21.setText(str(conf_intveral1))
            self.lineEdit_25.setText(str(''))
            if self.radioButton_3.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist3( N1,N2, x, y)
            elif self.radioButton.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist2(N1, x, mean1, std1, m1, n1)
            elif self.radioButton_5.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist2(N2, y, mean2, std2, m2, n2)
                
        elif self.comboBox.currentText() == "均值差（方差未知）":
            m1=float(self.doubleSpinBox_2.text())
            n1=float(self.doubleSpinBox_3.text())
            m2=float(self.doubleSpinBox_4.text())
            n2=float(self.doubleSpinBox_5.text())
            if self.data == None:
                N1=int(self.spinBox.text())
                x = np.random.normal(m1, n1, N1)
                N2=int(self.spinBox_2.text())
                y = np.random.normal(m2, n2, N2)
            else:
                x = np.array(self.data[0])
                y = np.array(self.data[1])
                N1 = len(x)
                N2 = len(y)
            mean1, std1 = x.mean(), x.std(ddof=1)
            self.lineEdit_3.setText(str(round(mean1, 3)))
            self.lineEdit_4.setText(str(round(std1, 3)))
            mean2, std2 = y.mean(), y.std(ddof=1)
            self.lineEdit_5.setText(str(round(mean2, 3)))
            self.lineEdit_6.setText(str(round(std2, 3)))
            a=float(self.doubleSpinBox.text())
            p=f.ppf(1-a/2,N1-1, N2-1)
            q=f.ppf(a/2,N1-1, N2-1)
            v=std1/(std2*q)
            w=std1/(std2*p)
            conf_intveral1=(round(v, 3), round(w, 3))
            self.lineEdit_25.setText(str(conf_intveral1))
            self.lineEdit_21.setText(str(''))
            if self.radioButton_3.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist3( N1,N2, x, y)
            elif self.radioButton.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist2(N1, x, mean1, std1, m1, n1)
            elif self.radioButton_5.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist2(N2, y, mean2, std2, m2, n2)
                
        elif self.comboBox.currentText() == "方差比":
            m1=float(self.doubleSpinBox_2.text())
            n1=float(self.doubleSpinBox_3.text())
            m2=float(self.doubleSpinBox_4.text())
            n2=float(self.doubleSpinBox_5.text())
            if self.data == None:
                N1=int(self.spinBox.text())
                x = np.random.normal(m1, n1, N1)
                N2=int(self.spinBox_2.text())
                y = np.random.normal(m2, n2, N2)
            else:
                x = np.array(self.data[0])
                y = np.array(self.data[1])
                N1 = len(x)
                N2 = len(y)
            mean1, std1 = x.mean(), x.std(ddof=1)
            self.lineEdit_3.setText(str(round(mean1, 3)))
            self.lineEdit_4.setText(str(round(std1, 3)))
            mean2, std2 = y.mean(), y.std(ddof=1)
            self.lineEdit_5.setText(str(round(mean2, 3)))
            self.lineEdit_6.setText(str(round(std2, 3)))
            a=float(self.doubleSpinBox.text())
            p=t.ppf(1-a/2,N1+N2-2)
            s=((N1-1)*std1+(N2-1)*std2)/(N1+N2-2)
            e=mean1-mean2-p*s*((1/N1+1/N2)**0.5)
            f=mean1-mean2+p*s*((1/N1+1/N2)**0.5)
            conf_intveral1=(round(e, 3), round(f, 3))
            self.lineEdit_21.setText(str(conf_intveral1))
            self.lineEdit_25.setText(str(''))
            if self.radioButton_3.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist3( N1,N2, x, y)
            elif self.radioButton.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist2(N1, x, mean1, std1, m1, n1)
            elif self.radioButton_5.isChecked():
                self.widget.setVisible(True)
                self.widget.mpl.draw_hist2(N2, y, mean2, std2, m2, n2)
        # TODO: not implemented yet
    

    @pyqtSlot()
    def on_pushButton_7_clicked(self):
        """
        Slot documentation goes here.
        """
        try:
            a=self.lineEdit.text()
            book = load_workbook(filename=a + ".xlsx")
            sheet = book.get_sheet_by_name("Sheet1")
            data=[[],  []]
            row_num=1
            for i in range(1,  3):
                while True:
                    if sheet.cell(row=row_num, column=i).value == None:
                        break
                    data[i - 1].append(sheet.cell(row=row_num, column=i).value)
                    row_num = row_num + 1
                row_num = 1
        except:
            QMessageBox.information(self, "标题", "请输入正确路径！", QMessageBox.Cancel)
        else:
            QMessageBox.information(self, "标题", "导入成功！", QMessageBox.Cancel)
        self.data = data
        # TODO: not implemented yet
        
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = double_confidence_interval_Form()
    ui.show()
    sys.exit(app.exec_()) 
