# -*- coding: utf-8 -*-

"""
Module implementing Form.
"""
from PyQt5.QtGui import QIcon, QPixmap, QPainter
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QWidget, QApplication
from openpyxl import load_workbook
from Ui_FenbuhanshuNihe import Ui_FenbuhanshuNihe_Form
from scipy import stats

class FenbuhanshuNihe_Form(QWidget, Ui_FenbuhanshuNihe_Form):
    """
    Class documentation goes here.
    """
    def __init__(self, parent=None):
        """
        Constructor
        
        @param parent reference to the parent widget
        @type QWidget
        """
        super(FenbuhanshuNihe_Form, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('./image/icon.png')) 
    
    def paintEvent(self,event):
        painter = QPainter(self)
        pixmap = QPixmap("./image/background.png")
        #绘制窗口背景，平铺到整个窗口，随着窗口改变而改变
        painter.drawPixmap(self.rect(),pixmap) 
    
    @pyqtSlot()
    def on_pushButton_clicked(self):
        """
        Slot documentation goes here.
        """
        a=self.lineEdit.text()
        book = load_workbook(filename=a + ".xlsx")
        sheet = book.get_sheet_by_name("Sheet1")
        data=[]
        row_num=1
        while True:
            if sheet.cell(row=row_num, column=1).value == None:
                break
            data.append(sheet.cell(row=row_num, column=1).value)
            row_num = row_num + 1
            
        parameter=stats.shapiro(data) 
        
        if parameter[1]>0.05:
            c=1
        else:
            c=0
        self.lineEdit_2.setText(str(c))
        print(data)
    
    
    @pyqtSlot()
    def on_pushButton_2_clicked(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        raise NotImplementedError
        
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = FenbuhanshuNihe_Form()
    ui.show()
    sys.exit(app.exec_())
